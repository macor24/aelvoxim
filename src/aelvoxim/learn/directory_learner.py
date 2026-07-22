"""directory_learner — 知识库目录自动学习引擎

递归扫描指定目录，提取文件内容，存入 knowledge_entries。
支持增量扫描（SHA256 去重），多种文件格式解析。

用法:
    from aelvoxim.learn.directory_learner import scan_directory
    result = scan_directory("/mnt/c/知识库")
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

log = logging.getLogger("aelvoxim.directory_learner")

# ── 已学习的文件记录 ──
_KNOWN_FILES: set[str] = set()  # SHA256 set for fast dedup
_LEARNER_CONFIG_PATH: Optional[Path] = None
_SCAN_CONFIG: Dict[str, Any] = {}  # in-memory scan config


# ═══════════════════════════════════════════
# 工具: SHA256 文件哈希
# ═══════════════════════════════════════════


def _sha256(path: Path) -> str:
    """Compute SHA256 of a file, reading in chunks."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


# ═══════════════════════════════════════════
# 递归扫描
# ═══════════════════════════════════════════


def _is_supported_file(path: Path) -> bool:
    """Check if file type is supported."""
    ext = path.suffix.lower()
    return ext in (
        ".txt", ".md", ".rst",
        ".pdf",
        ".docx", ".doc",
        ".csv", ".xlsx", ".xls",
        ".html", ".htm",
        ".json", ".yaml", ".yml", ".toml",
        ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".webp",
    )


def _extract_text(path: Path) -> Tuple[str, str]:
    """Extract text content from a file based on its extension.

    Returns:
        (text_content, title)
    """
    ext = path.suffix.lower()
    title = path.stem

    try:
        # Text formats
        if ext in (".txt", ".md", ".rst"):
            text = path.read_text(encoding="utf-8", errors="replace")
            return text, title

        # HTML
        if ext in (".html", ".htm"):
            try:
                from lxml import html as lh
                tree = lh.parse(str(path))
                text = tree.text_content().strip()
                if not text:
                    text = path.read_text(encoding="utf-8", errors="replace")
                return text, title
            except ImportError:
                text = path.read_text(encoding="utf-8", errors="replace")
                return text, title

        # PDF
        if ext == ".pdf":
            try:
                import pymupdf  # PyMuPDF
                doc = pymupdf.open(str(path))
                text = "\n".join(page.get_text() for page in doc)
                doc.close()
                if len(text.strip()) > 50:
                    return text, title
                # If too little text, fall through to OCR
            except ImportError:
                pass
            # OCR fallback for scanned PDFs
            try:
                from aelvoxim.learn.ocr_extract import ocr_extract
                result = ocr_extract(path)
                if result:
                    return result, title
            except ImportError:
                pass
            return "", title

        # DOCX
        if ext in (".docx", ".doc"):
            try:
                from docx import Document
                doc = Document(str(path))
                text = "\n".join(p.text for p in doc.paragraphs)
                return text, title
            except ImportError:
                return "", title

        # CSV
        if ext == ".csv":
            try:
                import csv as _csv
                with open(path, newline="", encoding="utf-8", errors="replace") as f:
                    reader = _csv.reader(f)
                    rows = []
                    for i, row in enumerate(reader):
                        if i > 100:
                            rows.append("... (truncated)")
                            break
                        rows.append(", ".join(row))
                    return "\n".join(rows), title
            except Exception:
                return path.read_text(encoding="utf-8", errors="replace"), title

        # Excel
        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i > 200:
                            break
                        rows.append(", ".join(str(c) if c is not None else "" for c in row))
                    sheets.append(f"=== {sheet_name} ===\n" + "\n".join(rows))
                wb.close()
                return "\n\n".join(sheets), title
            except ImportError:
                return "", title

        # JSON/YAML config
        if ext in (".json", ".yaml", ".yml", ".toml"):
            try:
                text = path.read_text(encoding="utf-8", errors="replace")
                return text, title
            except Exception:
                return "", title

        # Images — OCR
        if ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".webp"):
            try:
                from aelvoxim.learn.ocr_extract import ocr_extract
                result = ocr_extract(path)
                if result:
                    return result, title
            except ImportError:
                pass
            return "", title

    except Exception as e:
        log.warning("Failed to extract %s: %s", path, e)

    return "", title


def scan_directory(
    directory: str,
    recursive: bool = True,
    user_id: str = "",
    source: str = "server",
    dry_run: bool = False,
) -> Dict[str, Any]:
    """Scan a directory for new or changed files and learn their contents.

    Args:
        directory: Path to scan.
        recursive: Whether to scan subdirectories.
        user_id: User identifier for multi-tenant isolation.
        source: 'server' or 'gateway'.
        dry_run: If True, only report what would be learned.

    Returns:
        dict with keys: scanned, new, skipped, errors, entries
    """
    from ..storage.db import execute, fetch_dict, fetch_one, use_pg

    if not use_pg():
        return {"error": "PostgreSQL not available", "scanned": 0, "new": 0, "skipped": 0, "errors": 0}

    root = Path(directory).expanduser().resolve()
    if not root.is_dir():
        return {"error": f"Not a directory: {directory}", "scanned": 0, "new": 0, "skipped": 0, "errors": 0}

    # Load known SHA256s for fast dedup
    try:
        rows = fetch_dict(
            "SELECT sha256 FROM knowledge_files WHERE source = %s AND (user_id = %s OR user_id = '')",
            (source, user_id or ""),
        ) or []
        known_shas = {r["sha256"] for r in rows}
    except Exception:
        known_shas = set()

    # Walk directory
    new_files: List[Path] = []
    skipped = 0
    errors = 0
    scanned = 0

    if recursive:
        it = root.rglob("*")
    else:
        it = root.glob("*")

    for fp in it:
        if not fp.is_file():
            continue
        if not _is_supported_file(fp):
            continue
        # Skip hidden files
        if fp.name.startswith("."):
            continue
        scanned += 1

        try:
            sha = _sha256(fp)
        except Exception:
            errors += 1
            continue

        if sha in known_shas:
            skipped += 1
            continue

        new_files.append(fp)
        if dry_run:
            known_shas.add(sha)  # avoid double counting in dry_run
            continue

    if dry_run:
        return {
            "scanned": scanned,
            "new": len(new_files),
            "skipped": skipped,
            "errors": errors,
            "files": [str(f) for f in new_files],
        }

    # Process new files
    learned = 0
    for fp in new_files:
        try:
            _learn_file(fp, source, user_id)
            learned += 1
        except Exception as e:
            log.exception("Failed to learn %s: %s", fp, e)
            errors += 1

    return {
        "scanned": scanned,
        "new": learned,
        "skipped": skipped,
        "errors": errors,
    }


# ═══════════════════════════════════════════
# 单个文件学习
# ═══════════════════════════════════════════


def _learn_file(path: Path, source: str, user_id: str) -> bool:
    """Extract content from a file and store it as a knowledge entry."""
    from ..storage.db import execute, fetch_one

    text, title = _extract_text(path)
    if not text.strip():
        # Empty content — still record the file so we don't retry
        execute(
            """INSERT INTO knowledge_files (path, sha256, file_size, title, file_type, source, user_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s)
               ON CONFLICT DO NOTHING""",
            (str(path), _sha256(path), path.stat().st_size, title, path.suffix.lower().lstrip("."), source, user_id or ""),
        )
        return False

    # Determine if small file (direct index) or large file (LLM summary)
    is_small = len(text) < 500

    if is_small:
        # Small file: store full text directly
        content = text[:10000]  # cap at 10k chars
        topic = _auto_topic(title, text)
        tags = _auto_tags(title, text)
    else:
        # Large file: generate summary via LLM
        content, topic, tags = _llm_summarize(title, text)

    # Insert into knowledge_entries
    try:
        execute(
            """INSERT INTO knowledge_entries (topic, title, content, status, tags, source, created_at, updated_at)
               VALUES (%s, %s, %s, 'active', %s::jsonb, %s, NOW(), NOW())
               ON CONFLICT (topic, title) DO UPDATE SET content=EXCLUDED.content, tags=EXCLUDED.tags, updated_at=NOW()""",
            (topic[:200], title[:300], content[:10000], json.dumps(tags), source),
        )
        # Get the entry id
        row = fetch_one(
            "SELECT id FROM knowledge_entries WHERE topic = %s AND title = %s",
            (topic[:200], title[:300]),
        )
        entry_id = str(row[0]) if row else None

        # Record the file
        execute(
            """INSERT INTO knowledge_files (path, sha256, file_size, title, file_type, source, user_id, entry_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
               ON CONFLICT DO NOTHING""",
            (str(path), _sha256(path), path.stat().st_size, title, path.suffix.lower().lstrip("."), source, user_id or "", entry_id),
        )
        return True
    except Exception as e:
        log.error("Failed to insert knowledge entry for %s: %s", path, e)
        return False


# ═══════════════════════════════════════════
# LLM 摘要生成
# ═══════════════════════════════════════════


def _auto_topic(title: str, text: str) -> str:
    """Auto-detect topic from title or first line."""
    title_stripped = title.strip()
    if title_stripped:
        return title_stripped[:200]
    first_line = text.split("\n")[0].strip().strip("#").strip()
    return first_line[:200] if first_line else "general"


def _auto_tags(title: str, text: str) -> List[str]:
    """Simple tag extraction from title words."""
    import re
    words = re.findall(r'[a-zA-Z\u4e00-\u9fff]{2,}', title)
    return [w.lower() for w in words[:10]]


def _llm_summarize(title: str, text: str) -> Tuple[str, str, List[str]]:
    """Generate summary and tags for a large file via LLM."""
    from ..learn.llm import call_llm
    from ..cortex.model_router import get_chat_model_config

    # Truncate text to avoid token overflow
    text_preview = text[:8000]

    prompt = (
        f"请分析以下文档，返回 JSON（不要用 markdown 代码块，只返回原始 JSON）：\n\n"
        f"文档标题: {title}\n\n"
        f"文档内容（截取）:\n{text_preview}\n\n"
        f"请返回以下 JSON 格式:\n"
        f'{{"summary": "200-500字的摘要", "topic": "主题分类", "tags": ["标签1", "标签2", ...]}}'
    )

    try:
        mc = get_chat_model_config()
        result = call_llm(mc, "", prompt, temperature=0.3, max_tokens=1000) or ""
        # Clean response
        result = result.strip().strip("`").strip()
        if result.startswith("json"):
            result = result[4:].strip()
        parsed = json.loads(result)
        summary = parsed.get("summary", text[:500])
        topic = parsed.get("topic", _auto_topic(title, text))
        tags = parsed.get("tags", _auto_tags(title, text))
        return summary, topic[:200], tags[:20]
    except Exception as e:
        log.warning("LLM summarization failed for %s: %s", title, e)
        # Fallback: use first 500 chars
        return text[:500], _auto_topic(title, text), _auto_tags(title, text)


# ═══════════════════════════════════════════
# 扫描配置持久化
# ═══════════════════════════════════════════


def _get_config_path() -> Path:
    """Get path to scan config file."""
    global _LEARNER_CONFIG_PATH
    if _LEARNER_CONFIG_PATH is None:
        from ..utils import DATA_DIR
        _LEARNER_CONFIG_PATH = DATA_DIR / "learn_directory_config.json"
    return _LEARNER_CONFIG_PATH


def load_config() -> Dict[str, Any]:
    """Load scan configuration."""
    global _SCAN_CONFIG
    cfg_path = _get_config_path()
    if cfg_path.exists():
        try:
            _SCAN_CONFIG = json.loads(cfg_path.read_text())
        except Exception:
            _SCAN_CONFIG = {}
    return _SCAN_CONFIG


def save_config(config: Dict[str, Any]) -> None:
    """Save scan configuration."""
    global _SCAN_CONFIG
    _SCAN_CONFIG = config
    cfg_path = _get_config_path()
    cfg_path.parent.mkdir(parents=True, exist_ok=True)
    cfg_path.write_text(json.dumps(config, indent=2, ensure_ascii=False))


def get_config() -> Dict[str, Any]:
    """Get current scan configuration."""
    return _SCAN_CONFIG or load_config()


# ═══════════════════════════════════════════
# 定时扫描任务
# ═══════════════════════════════════════════


def run_scheduled_scan() -> Dict[str, Any]:
    """Run a scan based on current configuration. Called by cron."""
    cfg = get_config()
    directory = cfg.get("directory", "")
    if not directory:
        return {"error": "No directory configured", "scanned": 0}

    recursive = cfg.get("recursive", True)
    source = cfg.get("source", "server")
    user_id = cfg.get("user_id", "")

    log.info("Scheduled scan: %s (recursive=%s, source=%s)", directory, recursive, source)
    result = scan_directory(directory, recursive=recursive, user_id=user_id, source=source)
    log.info("Scheduled scan result: %s", result)
    return result
